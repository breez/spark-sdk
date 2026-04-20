#!/usr/bin/env python3
"""Build project analytics Excel from cached GitHub PR JSON and local git log.

Inputs (all under scripts/analytics/data/):
  - prs_all.json         List of PRs from GitHub `list_pull_requests` (merged pages).
  - git_log_main.txt     Output of `git log main --numstat --format='__COMMIT__%H\t%aI\t%an\t%ae\t%s'`.
  - pr_enrichment.json   (optional) Map of PR number -> {additions, deletions, changed_files,
                         commits, first_review_at, review_count, comment_count, requested_reviewers,
                         labels}. Populated by enrich_prs.py.

Output:
  - scripts/analytics/project_analytics.xlsx

Usage:
  python3 scripts/analytics/build_analytics.py
"""

from __future__ import annotations

import json
import os
import statistics
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"
OUT = ROOT / "project_analytics.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def parse_iso(s: str | None) -> datetime | None:
    if not s:
        return None
    dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
    # Excel doesn't allow tz-aware datetimes; normalise to naive UTC.
    if dt.tzinfo is not None:
        dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt


def iso_week(dt: datetime) -> str:
    y, w, _ = dt.isocalendar()
    return f"{y}-W{w:02d}"


def month_key(dt: datetime) -> str:
    return dt.strftime("%Y-%m")


def hours_between(a: datetime | None, b: datetime | None) -> float | None:
    if a is None or b is None:
        return None
    return round((b - a).total_seconds() / 3600.0, 2)


def write_header(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def autosize(ws, max_width: int = 60) -> None:
    for col_idx, col in enumerate(ws.columns, 1):
        longest = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            longest = max(longest, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(longest + 2, 10), max_width)


# ---------- Loaders ----------

def load_prs() -> list[dict]:
    with open(DATA / "prs_all.json") as f:
        return json.load(f)


def load_enrichment() -> dict[int, dict]:
    p = DATA / "pr_enrichment.json"
    if not p.exists():
        return {}
    with open(p) as f:
        return {int(k): v for k, v in json.load(f).items()}


def load_git_log() -> list[dict]:
    """Parse the numstat git log dump into a list of commits with per-file stats."""
    commits: list[dict] = []
    with open(DATA / "git_log_main.txt") as f:
        data = f.read()
    # Split on the marker; first chunk is empty.
    for chunk in data.split("__COMMIT__"):
        chunk = chunk.strip("\n")
        if not chunk:
            continue
        lines = chunk.split("\n")
        header = lines[0]
        parts = header.split("\t")
        if len(parts) < 5:
            continue
        sha, iso, name, email, subject = parts[0], parts[1], parts[2], parts[3], "\t".join(parts[4:])
        additions = 0
        deletions = 0
        files = 0
        for ln in lines[1:]:
            ln = ln.strip()
            if not ln:
                continue
            bits = ln.split("\t")
            if len(bits) < 3:
                continue
            a, d = bits[0], bits[1]
            # Binary files show up as '-'
            if a != "-":
                try:
                    additions += int(a)
                except ValueError:
                    pass
            if d != "-":
                try:
                    deletions += int(d)
                except ValueError:
                    pass
            files += 1
        commits.append({
            "sha": sha,
            "date": parse_iso(iso),
            "author": name,
            "email": email,
            "subject": subject,
            "additions": additions,
            "deletions": deletions,
            "files": files,
            "is_merge": subject.startswith("Merge "),
        })
    return commits


# ---------- Sheet builders ----------

def sheet_prs(wb: Workbook, prs: list[dict], enrich: dict[int, dict]) -> None:
    ws = wb.create_sheet("PRs")
    headers = [
        "number", "title", "author", "state", "merged", "draft",
        "created_at", "merged_at", "closed_at", "updated_at",
        "lead_time_hours", "close_time_hours",
        "base_ref", "head_ref", "html_url",
        "created_week", "created_month",
        "additions", "deletions", "changed_files", "commits",
        "review_count", "comment_count", "first_review_hours",
        "requested_reviewers", "labels",
    ]
    write_header(ws, headers)
    for pr in prs:
        created = parse_iso(pr.get("created_at"))
        merged = parse_iso(pr.get("merged_at"))
        closed = parse_iso(pr.get("closed_at"))
        updated = parse_iso(pr.get("updated_at"))
        e = enrich.get(pr["number"], {})
        first_review = parse_iso(e.get("first_review_at"))
        row = [
            pr["number"],
            pr.get("title", ""),
            (pr.get("user") or {}).get("login", ""),
            pr.get("state"),
            bool(pr.get("merged_at")),
            bool(pr.get("draft")),
            created,
            merged,
            closed,
            updated,
            hours_between(created, merged),
            hours_between(created, closed),
            (pr.get("base") or {}).get("ref", ""),
            (pr.get("head") or {}).get("ref", ""),
            pr.get("html_url", f"https://github.com/breez/spark-sdk/pull/{pr['number']}"),
            iso_week(created) if created else None,
            month_key(created) if created else None,
            e.get("additions"),
            e.get("deletions"),
            e.get("changed_files"),
            e.get("commits"),
            e.get("review_count"),
            e.get("comment_count"),
            hours_between(created, first_review),
            ", ".join(e.get("requested_reviewers", []) or []),
            ", ".join(e.get("labels", []) or []),
        ]
        ws.append(row)
    # Format date columns
    date_cols = [headers.index(c) + 1 for c in ("created_at", "merged_at", "closed_at", "updated_at")]
    for r in range(2, ws.max_row + 1):
        for c in date_cols:
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                cell.number_format = "yyyy-mm-dd hh:mm"
    autosize(ws)


def sheet_weekly_velocity(wb: Workbook, prs: list[dict], commits: list[dict]) -> None:
    ws = wb.create_sheet("Weekly Velocity")
    headers = ["week", "prs_opened", "prs_merged", "prs_closed_unmerged",
               "commits_to_main", "commits_additions", "commits_deletions",
               "unique_pr_authors", "unique_commit_authors", "cumulative_merged"]
    write_header(ws, headers)
    opened = Counter()
    merged = Counter()
    closed_unmerged = Counter()
    pr_authors: dict[str, set[str]] = defaultdict(set)
    commit_by_week = Counter()
    add_by_week = Counter()
    del_by_week = Counter()
    commit_authors: dict[str, set[str]] = defaultdict(set)

    for pr in prs:
        created = parse_iso(pr.get("created_at"))
        if created:
            w = iso_week(created)
            opened[w] += 1
            pr_authors[w].add((pr.get("user") or {}).get("login", ""))
        m = parse_iso(pr.get("merged_at"))
        if m:
            merged[iso_week(m)] += 1
        elif pr.get("state") == "closed":
            c = parse_iso(pr.get("closed_at"))
            if c:
                closed_unmerged[iso_week(c)] += 1

    for c in commits:
        if c["is_merge"]:
            continue  # exclude merge commits from counts
        w = iso_week(c["date"])
        commit_by_week[w] += 1
        add_by_week[w] += c["additions"]
        del_by_week[w] += c["deletions"]
        commit_authors[w].add(c["author"])

    all_weeks = sorted(set(opened) | set(merged) | set(closed_unmerged) | set(commit_by_week))
    cum = 0
    for w in all_weeks:
        cum += merged[w]
        ws.append([
            w, opened[w], merged[w], closed_unmerged[w],
            commit_by_week[w], add_by_week[w], del_by_week[w],
            len(pr_authors[w]), len(commit_authors[w]), cum,
        ])
    autosize(ws)


def sheet_monthly_velocity(wb: Workbook, prs: list[dict], commits: list[dict]) -> None:
    ws = wb.create_sheet("Monthly Velocity")
    headers = ["month", "prs_opened", "prs_merged", "prs_closed_unmerged",
               "commits_to_main", "commits_additions", "commits_deletions",
               "unique_pr_authors", "unique_commit_authors",
               "avg_lead_time_hours", "median_lead_time_hours"]
    write_header(ws, headers)

    opened = Counter()
    merged = Counter()
    closed_unmerged = Counter()
    pr_authors: dict[str, set[str]] = defaultdict(set)
    commit_by_m = Counter()
    add_by_m = Counter()
    del_by_m = Counter()
    commit_authors: dict[str, set[str]] = defaultdict(set)
    lead_times: dict[str, list[float]] = defaultdict(list)

    for pr in prs:
        created = parse_iso(pr.get("created_at"))
        if created:
            mk = month_key(created)
            opened[mk] += 1
            pr_authors[mk].add((pr.get("user") or {}).get("login", ""))
        m = parse_iso(pr.get("merged_at"))
        if m and created:
            mk = month_key(m)
            merged[mk] += 1
            lead_times[mk].append((m - created).total_seconds() / 3600.0)
        elif pr.get("state") == "closed":
            c = parse_iso(pr.get("closed_at"))
            if c:
                closed_unmerged[month_key(c)] += 1

    for c in commits:
        if c["is_merge"]:
            continue
        mk = month_key(c["date"])
        commit_by_m[mk] += 1
        add_by_m[mk] += c["additions"]
        del_by_m[mk] += c["deletions"]
        commit_authors[mk].add(c["author"])

    all_m = sorted(set(opened) | set(merged) | set(closed_unmerged) | set(commit_by_m))
    for mk in all_m:
        lt = lead_times[mk]
        ws.append([
            mk, opened[mk], merged[mk], closed_unmerged[mk],
            commit_by_m[mk], add_by_m[mk], del_by_m[mk],
            len(pr_authors[mk]), len(commit_authors[mk]),
            round(statistics.mean(lt), 2) if lt else None,
            round(statistics.median(lt), 2) if lt else None,
        ])
    autosize(ws)


def sheet_author_stats(wb: Workbook, prs: list[dict], commits: list[dict], enrich: dict[int, dict]) -> None:
    ws = wb.create_sheet("Author Stats")
    headers = ["author", "prs_opened", "prs_merged", "merge_rate",
               "avg_lead_time_hours", "median_lead_time_hours",
               "total_additions", "total_deletions", "total_commits_to_main",
               "first_pr", "last_pr"]
    write_header(ws, headers)
    by_author: dict[str, dict] = defaultdict(lambda: {
        "opened": 0, "merged": 0, "lead_times": [], "first": None, "last": None,
        "additions": 0, "deletions": 0, "commits": 0,
    })
    for pr in prs:
        a = (pr.get("user") or {}).get("login", "")
        if not a:
            continue
        created = parse_iso(pr.get("created_at"))
        merged = parse_iso(pr.get("merged_at"))
        d = by_author[a]
        d["opened"] += 1
        if created:
            d["first"] = created if d["first"] is None or created < d["first"] else d["first"]
            d["last"] = created if d["last"] is None or created > d["last"] else d["last"]
        if merged:
            d["merged"] += 1
            if created:
                d["lead_times"].append((merged - created).total_seconds() / 3600.0)
        # Enrichment adds (additions, deletions only for merged usually)
        e = enrich.get(pr["number"])
        if e and merged:
            d["additions"] += e.get("additions") or 0
            d["deletions"] += e.get("deletions") or 0

    # Aggregate commits to main per author (by name — GH login may differ).
    commits_by_author = Counter()
    for c in commits:
        if c["is_merge"]:
            continue
        commits_by_author[c["author"]] += 1

    # We don't have a robust name->login mapping. Show per GitHub login for PR stats,
    # and attach commit counts via a best-effort name match below.
    all_authors = sorted(by_author.keys())
    for a in all_authors:
        d = by_author[a]
        lt = d["lead_times"]
        ws.append([
            a, d["opened"], d["merged"],
            round(d["merged"] / d["opened"], 3) if d["opened"] else None,
            round(statistics.mean(lt), 2) if lt else None,
            round(statistics.median(lt), 2) if lt else None,
            d["additions"] or None,
            d["deletions"] or None,
            None,  # commits column left blank (see Git Authors sheet)
            d["first"], d["last"],
        ])
    # Date cols
    for r in range(2, ws.max_row + 1):
        for col_name in ("first_pr", "last_pr"):
            c = ws.cell(row=r, column=headers.index(col_name) + 1)
            if c.value is not None:
                c.number_format = "yyyy-mm-dd"
    autosize(ws)


def sheet_git_authors(wb: Workbook, commits: list[dict]) -> None:
    ws = wb.create_sheet("Git Authors")
    headers = ["author", "email_sample", "commits_to_main",
               "additions", "deletions", "files_touched",
               "first_commit", "last_commit"]
    write_header(ws, headers)
    agg: dict[str, dict] = defaultdict(lambda: {
        "count": 0, "add": 0, "del": 0, "files": 0,
        "first": None, "last": None, "email": None,
    })
    for c in commits:
        if c["is_merge"]:
            continue
        a = agg[c["author"]]
        a["count"] += 1
        a["add"] += c["additions"]
        a["del"] += c["deletions"]
        a["files"] += c["files"]
        a["email"] = a["email"] or c["email"]
        if a["first"] is None or c["date"] < a["first"]:
            a["first"] = c["date"]
        if a["last"] is None or c["date"] > a["last"]:
            a["last"] = c["date"]
    for name in sorted(agg, key=lambda k: -agg[k]["count"]):
        d = agg[name]
        ws.append([name, d["email"], d["count"], d["add"], d["del"], d["files"],
                   d["first"], d["last"]])
    for r in range(2, ws.max_row + 1):
        for col_name in ("first_commit", "last_commit"):
            c = ws.cell(row=r, column=headers.index(col_name) + 1)
            if c.value is not None:
                c.number_format = "yyyy-mm-dd hh:mm"
    autosize(ws)


def sheet_commits(wb: Workbook, commits: list[dict]) -> None:
    ws = wb.create_sheet("Commits")
    headers = ["sha", "date", "author", "email", "subject",
               "additions", "deletions", "files", "is_merge", "week", "month"]
    write_header(ws, headers)
    for c in commits:
        ws.append([
            c["sha"][:12], c["date"], c["author"], c["email"], c["subject"],
            c["additions"], c["deletions"], c["files"], c["is_merge"],
            iso_week(c["date"]) if c["date"] else None,
            month_key(c["date"]) if c["date"] else None,
        ])
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=headers.index("date") + 1)
        if cell.value is not None:
            cell.number_format = "yyyy-mm-dd hh:mm"
    autosize(ws, max_width=80)


def sheet_summary(wb: Workbook, prs: list[dict], commits: list[dict], enrich: dict[int, dict]) -> None:
    ws = wb.create_sheet("Summary", 0)  # first sheet
    ws.append(["Breez Spark SDK — Project Analytics"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=16)
    ws.append([])

    merged_prs = [p for p in prs if p.get("merged_at")]
    open_prs = [p for p in prs if p.get("state") == "open"]
    closed_unmerged = [p for p in prs if p.get("state") == "closed" and not p.get("merged_at")]
    lead_times = []
    for p in merged_prs:
        a = parse_iso(p["created_at"])
        b = parse_iso(p["merged_at"])
        if a and b:
            lead_times.append((b - a).total_seconds() / 3600.0)

    non_merge = [c for c in commits if not c["is_merge"]]
    total_add = sum(c["additions"] for c in non_merge)
    total_del = sum(c["deletions"] for c in non_merge)

    first_pr = min((parse_iso(p["created_at"]) for p in prs if p.get("created_at")), default=None)
    last_pr = max((parse_iso(p["created_at"]) for p in prs if p.get("created_at")), default=None)
    first_commit = min((c["date"] for c in commits if c["date"]), default=None)
    last_commit = max((c["date"] for c in commits if c["date"]), default=None)

    rows = [
        ("Date range (PRs)", f"{first_pr:%Y-%m-%d} → {last_pr:%Y-%m-%d}" if first_pr and last_pr else "n/a"),
        ("Date range (commits on main)", f"{first_commit:%Y-%m-%d} → {last_commit:%Y-%m-%d}" if first_commit and last_commit else "n/a"),
        ("", ""),
        ("Total PRs", len(prs)),
        ("  merged", len(merged_prs)),
        ("  open", len(open_prs)),
        ("  closed (not merged)", len(closed_unmerged)),
        ("Merge rate", f"{len(merged_prs) / len(prs):.1%}" if prs else "n/a"),
        ("Unique PR authors", len({(p.get('user') or {}).get('login') for p in prs if p.get('user')})),
        ("", ""),
        ("Avg lead time (h)", round(statistics.mean(lead_times), 1) if lead_times else "n/a"),
        ("Median lead time (h)", round(statistics.median(lead_times), 1) if lead_times else "n/a"),
        ("  min", round(min(lead_times), 1) if lead_times else "n/a"),
        ("  max", round(max(lead_times), 1) if lead_times else "n/a"),
        ("", ""),
        ("Commits to main (excl. merges)", len(non_merge)),
        ("  total additions", total_add),
        ("  total deletions", total_del),
        ("  net lines", total_add - total_del),
        ("Unique commit authors", len({c["author"] for c in non_merge})),
        ("", ""),
        ("PRs with review enrichment", sum(1 for p in prs if p["number"] in enrich)),
    ]
    for k, v in rows:
        ws.append([k, v])
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=1):
        for c in row:
            c.font = Font(bold=True)
    autosize(ws, max_width=50)


# ---------- Main ----------

def main() -> None:
    prs = load_prs()
    commits = load_git_log()
    enrich = load_enrichment()
    print(f"Loaded {len(prs)} PRs, {len(commits)} commits, {len(enrich)} enriched PRs.")

    wb = Workbook()
    # Remove default sheet; Summary inserts at index 0 below.
    default = wb.active
    wb.remove(default)

    sheet_summary(wb, prs, commits, enrich)
    sheet_prs(wb, prs, enrich)
    sheet_weekly_velocity(wb, prs, commits)
    sheet_monthly_velocity(wb, prs, commits)
    sheet_author_stats(wb, prs, commits, enrich)
    sheet_git_authors(wb, commits)
    sheet_commits(wb, commits)

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
